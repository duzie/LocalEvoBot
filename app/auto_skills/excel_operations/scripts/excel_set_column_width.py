import openpyxl
import os

def excel_set_column_width(file_path: str, column_letter: str, width: float, sheet_name: str = "Sheet1") -> str:
    """
    设置Excel列宽
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        column_letter: 列字母，如 'A'、'B'、'C'
        width: 列宽值
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当列字母无效或宽度值无效时
    """
    try:
        # 检查宽度值
        if width <= 0:
            raise ValueError(f"列宽必须大于0，当前值: {width}")
        
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 设置列宽
        sheet.column_dimensions[column_letter].width = width
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功设置列 {column_letter} 的宽度为 {width}"
        
    except Exception as e:
        raise Exception(f"设置Excel列宽失败: {str(e)}")