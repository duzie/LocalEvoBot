import openpyxl
import os
from typing import List, Dict, Any

def excel_find_value(file_path: str, search_value: str, sheet_name: str = "Sheet1", exact_match: bool = False) -> List[Dict[str, Any]]:
    """
    在Excel文件中查找包含指定值的单元格
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        search_value: 要查找的值
        sheet_name: 工作表名称，默认为 'Sheet1'
        exact_match: 是否精确匹配，默认为False（包含匹配）
        
    Returns:
        包含查找结果的列表，每个结果包含单元格地址和值
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当工作表不存在时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 查找匹配的单元格
        results = []
        
        for row in sheet.iter_rows():
            for cell in row:
                cell_value = cell.value
                
                if cell_value is not None:
                    # 转换为字符串进行比较
                    str_value = str(cell_value)
                    
                    if exact_match:
                        # 精确匹配
                        if str_value == search_value:
                            results.append({
                                "cell_address": cell.coordinate,
                                "value": cell_value,
                                "row": cell.row,
                                "column": cell.column,
                                "column_letter": cell.column_letter
                            })
                    else:
                        # 包含匹配
                        if search_value.lower() in str_value.lower():
                            results.append({
                                "cell_address": cell.coordinate,
                                "value": cell_value,
                                "row": cell.row,
                                "column": cell.column,
                                "column_letter": cell.column_letter
                            })
        
        # 关闭工作簿
        workbook.close()
        
        return results
        
    except Exception as e:
        raise Exception(f"查找Excel值失败: {str(e)}")