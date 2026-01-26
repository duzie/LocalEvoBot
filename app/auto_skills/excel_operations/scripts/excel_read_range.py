import openpyxl
import os
from typing import List, Any

def excel_read_range(file_path: str, range_address: str, sheet_name: str = "Sheet1") -> List[List[Any]]:
    """
    读取Excel文件中指定范围的数据
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        range_address: 范围地址，如 'A1:C10'、'B2:D5'
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        二维列表，包含范围内的所有数据
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当范围地址无效时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 解析范围地址
        if ':' not in range_address:
            raise ValueError(f"无效的范围地址格式: {range_address}。请使用格式如 'A1:C10'")
        
        start_cell, end_cell = range_address.split(':')
        
        # 获取范围数据
        data = []
        for row in sheet[start_cell:end_cell]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        
        # 关闭工作簿
        workbook.close()
        
        return data
        
    except Exception as e:
        raise Exception(f"读取Excel范围数据失败: {str(e)}")