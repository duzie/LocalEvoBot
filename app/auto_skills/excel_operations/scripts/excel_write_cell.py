import openpyxl
import os
from typing import Optional, Any

def excel_write_cell(file_path: str, cell_address: str, value: Any, sheet_name: str = "Sheet1") -> str:
    """
    向Excel文件中指定单元格写入数据
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        cell_address: 单元格地址，如 'A1'、'B2'
        value: 要写入的值
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        ValueError: 当单元格地址无效时
    """
    try:
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 写入单元格值
        sheet[cell_address].value = value
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功写入单元格 {cell_address}: {value}"
        
    except Exception as e:
        raise Exception(f"写入Excel单元格失败: {str(e)}")