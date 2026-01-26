import openpyxl
import os
from typing import Dict

def excel_get_dimensions(file_path: str, sheet_name: str = "Sheet1") -> Dict[str, int]:
    """
    获取指定工作表的行数和列数
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        包含行数和列数的字典
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当工作表不存在时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 获取最大行和最大列
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        # 计算实际使用的行数和列数（排除空行空列）
        actual_rows = 0
        actual_cols = 0
        
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
            row_has_data = any(cell.value is not None for cell in row)
            if row_has_data:
                actual_rows += 1
                # 更新实际列数
                for col_idx, cell in enumerate(row):
                    if cell.value is not None:
                        actual_cols = max(actual_cols, col_idx + 1)
        
        # 关闭工作簿
        workbook.close()
        
        return {
            "max_rows": max_row,
            "max_columns": max_column,
            "actual_rows": actual_rows,
            "actual_columns": actual_cols
        }
        
    except Exception as e:
        raise Exception(f"获取Excel工作表维度失败: {str(e)}")