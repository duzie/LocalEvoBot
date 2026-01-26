import openpyxl
import os
from typing import Optional

def excel_create_sheet(file_path: str, sheet_name: str, position: Optional[int] = None) -> str:
    """
    在Excel文件中创建新的工作表
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        sheet_name: 新工作表的名称
        position: 插入位置（从0开始），默认为最后
        
    Returns:
        成功消息
        
    Raises:
        ValueError: 当工作表名称已存在时
    """
    try:
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            default_sheet = workbook.active
            default_sheet.title = "Sheet1"
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表名称是否已存在
        if sheet_name in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 已存在。可用工作表: {workbook.sheetnames}")
        
        # 创建新工作表
        if position is not None:
            # 在指定位置创建
            workbook.create_sheet(title=sheet_name, index=position)
        else:
            # 在最后创建
            workbook.create_sheet(title=sheet_name)
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        position_info = f"位置 {position}" if position is not None else "最后位置"
        return f"成功创建工作表 '{sheet_name}' ({position_info})"
        
    except Exception as e:
        raise Exception(f"创建Excel工作表失败: {str(e)}")