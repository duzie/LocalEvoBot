import openpyxl
import os

def excel_delete_sheet(file_path: str, sheet_name: str) -> str:
    """
    删除Excel文件中的指定工作表
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        sheet_name: 要删除的工作表名称
        
    Returns:
        成功消息
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当工作表不存在或尝试删除最后一个工作表时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 检查是否是最后一个工作表
        if len(workbook.sheetnames) == 1:
            raise ValueError("不能删除最后一个工作表")
        
        # 删除工作表
        sheet_to_delete = workbook[sheet_name]
        workbook.remove(sheet_to_delete)
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功删除工作表 '{sheet_name}'"
        
    except Exception as e:
        raise Exception(f"删除Excel工作表失败: {str(e)}")