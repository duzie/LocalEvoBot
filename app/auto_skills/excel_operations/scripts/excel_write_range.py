import openpyxl
import os
from typing import List, Any

def excel_write_range(file_path: str, range_address: str, data: List[List[Any]], sheet_name: str = "Sheet1") -> str:
    """
    向Excel文件中指定范围写入数据
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        range_address: 范围地址，如 'A1:C10'、'B2:D5'
        data: 二维列表数据
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        ValueError: 当范围地址无效或数据维度不匹配时
    """
    try:
        # 检查数据是否为二维列表
        if not isinstance(data, list) or not all(isinstance(row, list) for row in data):
            raise ValueError("数据必须是二维列表")
        
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 解析范围地址
        if ':' not in range_address:
            raise ValueError(f"无效的范围地址格式: {range_address}。请使用格式如 'A1:C10'")
        
        start_cell, end_cell = range_address.split(':')
        
        # 获取范围的行列数
        start_col = openpyxl.utils.column_index_from_string(start_cell[0])
        start_row = int(start_cell[1:])
        end_col = openpyxl.utils.column_index_from_string(end_cell[0])
        end_row = int(end_cell[1:])
        
        # 计算范围的行列数
        range_rows = end_row - start_row + 1
        range_cols = end_col - start_col + 1
        
        # 检查数据维度是否匹配
        if len(data) != range_rows:
            raise ValueError(f"数据行数 ({len(data)}) 与范围行数 ({range_rows}) 不匹配")
        
        for i, row in enumerate(data):
            if len(row) != range_cols:
                raise ValueError(f"第 {i+1} 行数据列数 ({len(row)}) 与范围列数 ({range_cols}) 不匹配")
        
        # 写入数据
        for i in range(range_rows):
            for j in range(range_cols):
                cell_row = start_row + i
                cell_col = start_col + j
                cell_address = f"{openpyxl.utils.get_column_letter(cell_col)}{cell_row}"
                sheet[cell_address].value = data[i][j]
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功写入范围 {range_address}: {len(data)} 行 × {range_cols} 列 数据"
        
    except Exception as e:
        raise Exception(f"写入Excel范围数据失败: {str(e)}")