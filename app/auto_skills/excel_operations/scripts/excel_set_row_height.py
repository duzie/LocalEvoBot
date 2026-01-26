import openpyxl
import os

def excel_set_row_height(file_path: str, row_number: int, height: float, sheet_name: str = "Sheet1") -> str:
    """
    设置Excel行高
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        row_number: 行号，从1开始
        height: 行高值
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当行号无效或高度值无效时
    """
    try:
        # 检查行号
        if row_number < 1:
            raise ValueError(f"行号必须大于等于1，当前值: {row_number}")
        
        # 检查高度值
        if height <= 0:
            raise ValueError(f"行高必须大于0，当前值: {height}")
        
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 设置行高
        sheet.row_dimensions[row_number].height = height
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功设置行 {row_number} 的高度为 {height}"
        
    except Exception as e:
        raise Exception(f"设置Excel行高失败: {str(e)}")