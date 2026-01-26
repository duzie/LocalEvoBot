import openpyxl
import os
from typing import Optional

def excel_average_column(file_path: str, column_letter: str, start_row: int = 1, end_row: Optional[int] = None, sheet_name: str = "Sheet1") -> float:
    """
    计算指定列的平均值
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        column_letter: 列字母，如 'A'、'B'、'C'
        start_row: 起始行号，默认为1
        end_row: 结束行号，默认为最后一行
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        列的平均值，如果没有数值则返回0
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当工作表不存在或列无效时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 获取最大行数
        max_row = sheet.max_row
        
        # 设置结束行
        if end_row is None:
            end_row = max_row
        else:
            end_row = min(end_row, max_row)
        
        # 检查起始行和结束行
        if start_row < 1:
            start_row = 1
        if start_row > end_row:
            raise ValueError(f"起始行 ({start_row}) 不能大于结束行 ({end_row})")
        
        # 计算总和和计数
        total = 0.0
        count = 0
        
        for row in range(start_row, end_row + 1):
            cell_address = f"{column_letter}{row}"
            cell_value = sheet[cell_address].value
            
            if cell_value is not None:
                try:
                    # 尝试转换为数值
                    if isinstance(cell_value, (int, float)):
                        total += float(cell_value)
                        count += 1
                    elif isinstance(cell_value, str):
                        # 尝试从字符串解析数值
                        try:
                            num_value = float(cell_value)
                            total += num_value
                            count += 1
                        except ValueError:
                            # 忽略非数值字符串
                            pass
                except (ValueError, TypeError):
                    # 忽略无法转换的值
                    pass
        
        # 关闭工作簿
        workbook.close()
        
        # 计算平均值
        if count > 0:
            return total / count
        else:
            return 0.0
        
    except Exception as e:
        raise Exception(f"计算Excel列平均值失败: {str(e)}")