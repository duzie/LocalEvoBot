import openpyxl
import os
from typing import Optional, Any

def excel_read_cell(file_path: str, cell_address: str, sheet_name: str = "Sheet1") -> Any:
    """
    读取Excel文件中指定单元格的值
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        cell_address: 单元格地址，如 'A1'、'B2'
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        单元格的值，如果文件不存在或单元格为空则返回None
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当单元格地址无效时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 读取单元格值
        cell_value = sheet[cell_address].value
        
        # 关闭工作簿
        workbook.close()
        
        return cell_value
        
    except Exception as e:
        raise Exception(f"读取Excel单元格失败: {str(e)}")