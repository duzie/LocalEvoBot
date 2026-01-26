import openpyxl
import os
from typing import Optional

def excel_sort_data(file_path: str, range_address: str, sort_by: str, ascending: bool = True, sheet_name: str = "Sheet1") -> str:
    """
    对Excel数据进行排序
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        range_address: 要排序的范围地址，如 'A1:D100'
        sort_by: 按哪一列排序，如 'A'、'B'、'C'
        ascending: 是否升序排序，默认为True
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当范围地址无效或排序列不存在时
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 解析范围地址
        if ':' not in range_address:
            raise ValueError(f"无效的范围地址格式: {range_address}。请使用格式如 'A1:D100'")
        
        start_cell, end_cell = range_address.split(':')
        
        # 获取范围的行列数
        start_col = openpyxl.utils.column_index_from_string(start_cell[0])
        start_row = int(start_cell[1:])
        end_col = openpyxl.utils.column_index_from_string(end_cell[0])
        end_row = int(end_cell[1:])
        
        # 检查排序列是否在范围内
        sort_col_index = openpyxl.utils.column_index_from_string(sort_by)
        if sort_col_index < start_col or sort_col_index > end_col:
            raise ValueError(f"排序列 '{sort_by}' 不在范围 {range_address} 内")
        
        # 读取数据
        data = []
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell_address = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                row_data.append(sheet[cell_address].value)
            data.append(row_data)
        
        # 获取排序列的索引（相对于范围）
        sort_col_rel_index = sort_col_index - start_col
        
        # 对数据进行排序
        data.sort(key=lambda x: x[sort_col_rel_index] if x[sort_col_rel_index] is not None else "", reverse=not ascending)
        
        # 写回排序后的数据
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                cell_row = start_row + i
                cell_col = start_col + j
                cell_address = f"{openpyxl.utils.get_column_letter(cell_col)}{cell_row}"
                sheet[cell_address].value = cell_value
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        order = "升序" if ascending else "降序"
        return f"成功对范围 {range_address} 按列 {sort_by} 进行{order}排序"
        
    except Exception as e:
        raise Exception(f"Excel数据排序失败: {str(e)}")