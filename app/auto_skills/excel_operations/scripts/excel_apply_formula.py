import openpyxl
import os

def excel_apply_formula(file_path: str, cell_address: str, formula: str, sheet_name: str = "Sheet1") -> str:
    """
    向Excel单元格应用公式
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        cell_address: 单元格地址，如 'A1'、'B2'
        formula: Excel公式，如 '=SUM(A1:A10)'、'=AVERAGE(B2:B20)'
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        ValueError: 当单元格地址无效或公式格式错误时
    """
    try:
        # 检查公式是否以等号开头
        if not formula.startswith('='):
            formula = '=' + formula
        
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 应用公式
        sheet[cell_address].value = formula
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功在单元格 {cell_address} 应用公式: {formula}"
        
    except Exception as e:
        raise Exception(f"应用Excel公式失败: {str(e)}")