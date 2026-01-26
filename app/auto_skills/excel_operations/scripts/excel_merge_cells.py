import openpyxl
import os

def excel_merge_cells(file_path: str, range_address: str, sheet_name: str = "Sheet1") -> str:
    """
    合并Excel单元格
    
    Args:
        file_path: Excel文件路径 (.xlsx 或 .xls)
        range_address: 要合并的范围地址，如 'A1:C1'
        sheet_name: 工作表名称，默认为 'Sheet1'
        
    Returns:
        成功消息
        
    Raises:
        FileNotFoundError: 当文件不存在时
        ValueError: 当范围地址无效时
    """
    try:
        # 检查文件是否存在，不存在则创建新工作簿
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            # 重命名默认工作表
            if sheet_name != "Sheet":
                default_sheet = workbook.active
                default_sheet.title = sheet_name
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在，不存在则创建
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
        
        # 获取工作表
        sheet = workbook[sheet_name]
        
        # 检查范围地址格式
        if ':' not in range_address:
            raise ValueError(f"无效的范围地址格式: {range_address}。请使用格式如 'A1:C1'")
        
        # 合并单元格
        sheet.merge_cells(range_address)
        
        # 保存工作簿
        workbook.save(file_path)
        workbook.close()
        
        return f"成功合并单元格范围: {range_address}"
        
    except Exception as e:
        raise Exception(f"合并Excel单元格失败: {str(e)}")